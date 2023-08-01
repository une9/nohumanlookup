import pymysql
from openpyxl import load_workbook
# import sqlparse

# db connection
target_db = pymysql.connect(
    user='root', 
    passwd='mysql1234', 
    host='localhost', 
    db='nhl', 
    charset='utf8'
)

cursor = target_db.cursor(pymysql.cursors.DictCursor)
# query = "SELECT sysdate() FROM dual;"


def find_cell_right_of_text(sheet, target_text):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column == 3 and target_text == row[1].value:
                print(cell.value)
                return cell.value
    return None


# read excel and get Query
load_wb = load_workbook("./sample_document.xlsx", data_only=True)
load_ws = load_wb['Sheet1']
target_text = 'sample'

query = find_cell_right_of_text(load_ws, target_text)
if query is None:
    print(f"'{target_text}' not found in the sheet.")

cursor.execute(query)
result = cursor.fetchall()

print("result: ", result)

def extract_table_names(query):
    table_names = []
    parsed_query = sqlparse.parse(query)

    for statement in parsed_query:
        for token in statement.tokens:
            if token.ttype is sqlparse.tokens.DML and token.value.upper() == 'FROM':
                table_names.extend(get_table_names_after_from(token))

    return table_names

def get_table_names_after_from(from_token):
    table_names = []
    for item in from_token.get_sublists():
        for item_token in item.flatten():
            if item_token.ttype == sqlparse.tokens.Name:
                table_names.append(item_token.value)
    return table_names

table_names = extract_table_names(query)
print(table_names)